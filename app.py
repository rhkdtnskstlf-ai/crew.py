import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ==============================
# 상단 UI 및 안내 문구
# ==============================
st.set_page_config(page_title="Crew Stay Generator", layout="wide")
st.title("🏨 Crew Hotel Stay Generator")

with st.expander("✅ 업데이트 내역 및 사용 가이드 (클릭하여 확인)", expanded=True):
    st.markdown("""
    ### 🚀 주요 업데이트 사항
    * **필터링 행 완전 제외**: 엑셀에서 **필터로 숨기거나 행 높이를 0으로 줄인 데이터**는 분석에서 완벽하게 제외됩니다.
    * **병합 셀 에러 해결**: 날짜 구분선(병합 셀) 생성 시 발생하던 열 너비 조정 오류를 수정했습니다.
    * **시간 정보 유지**: 엑셀 결과물에 체크인/아웃 시간이 정상적으로 표시됩니다.
    * **베이징 레이오버 처리 개선**: 레이오버 편(예: KE2061) 출발 다음날 아침 복귀편(예: KE2062)을 정확히 삭제합니다.
    * **파란색 하이라이트 수정**: 레이오버 stay 행만 파란색으로 표시되며, 같은 크루의 다른 stay 행은 영향받지 않습니다.

    ### 🔍 오류 발생 시 체크리스트
    1. **컬럼명 확인**: `Crew ID`, `Crew Name`, `Rank`, `Arr Flt`, `Arr Flt Date`, `Dep Flt`, `Dep Flt Date` 컬럼이 정확히 있는지 확인하세요.
    2. **날짜 형식**: 날짜 컬럼에 텍스트나 깨진 값이 섞여 있으면 해당 행은 자동으로 제외됩니다.
    3. **파일 닫기**: 엑셀 파일이 다른 프로그램에서 열려 있으면 업로드 오류가 발생할 수 있습니다.
    """)

st.info("💡 원본 엑셀에서 필요한 데이터만 필터링한 후 업로드하시면 더욱 정확합니다.")

# ==============================
# 베이징 레이오버 출발편 입력 UI
# ==============================
st.subheader("✈️ 베이징 레이오버 출발편 설정")
st.markdown("해당 편으로 **출발(Dep Flt)하는 stay**만 Checkin 기준 **3박으로 고정**됩니다.")

layover_input = st.text_input(
    "레이오버 출발편 번호 입력 (여러 편수는 쉼표로 구분, 예: KE2061, KE2063)",
    placeholder="KE2061, KE2063"
)

layover_flights = []
if layover_input.strip():
    layover_flights = [f.strip().upper() for f in layover_input.split(",") if f.strip()]

if layover_flights:
    st.success(f"등록된 레이오버 출발편: **{', '.join(layover_flights)}**")

uploaded_file = st.file_uploader("엑셀 파일 업로드", type=["xlsx"])

# ==============================
# 숨겨진 행을 제외하고 읽어오는 함수
# ==============================
def read_excel_visible_only(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    data = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        row_idx = row[0].row
        if ws.row_dimensions[row_idx].hidden == False and ws.row_dimensions[row_idx].height != 0:
            data.append([cell.value for cell in row])
    return pd.DataFrame(data, columns=headers)

# ==============================
# 파일 처리 시작
# ==============================
if uploaded_file:
    df = read_excel_visible_only(uploaded_file)

    if df.empty:
        st.error("데이터가 없거나 모든 행이 숨겨져 있습니다.")
        st.stop()

    df.columns = df.columns.str.strip()

    required_cols = ['Crew ID', 'Crew Name', 'Rank', 'Arr Flt', 'Arr Flt Date', 'Dep Flt', 'Dep Flt Date']
    df = df[required_cols].copy()

    df['Arr Flt Date'] = pd.to_datetime(df['Arr Flt Date'], dayfirst=True, errors='coerce')
    df['Dep Flt Date'] = pd.to_datetime(df['Dep Flt Date'], dayfirst=True, errors='coerce')
    df = df.dropna(subset=['Crew ID', 'Arr Flt Date', 'Dep Flt Date'])

    df = df.sort_values(['Crew ID', 'Arr Flt Date'])

    # ==============================
    # 베이징 레이오버 처리:
    # 레이오버 출발편(예: KE2061)으로 출발하는 행의 다음날(+1일) 아침 복귀 행 삭제
    # 이유: KE2062처럼 복귀편이 레이오버 출발 다음날 새벽에 도착하므로 +1일이 정확함
    # ==============================
    if layover_flights:
        dep_flt_upper_raw = df['Dep Flt'].astype(str).str.strip().str.upper()
        layover_mask_raw  = dep_flt_upper_raw.isin(layover_flights)
        layover_rows_raw  = df[layover_mask_raw]

        # (crew_id, 삭제할 날짜) 쌍 수집: 레이오버 출발일 +1일 (복귀편이 다음날 아침)
        drop_keys = set()
        for _, row in layover_rows_raw.iterrows():
            crew_id  = row['Crew ID']
            dep_date = row['Dep Flt Date'].date()
            target_date = dep_date + pd.Timedelta(days=1)  # ✅ +2일 → +1일로 수정
            drop_keys.add((crew_id, target_date))

        def should_drop(row):
            return (row['Crew ID'], row['Arr Flt Date'].date()) in drop_keys

        df = df[~df.apply(should_drop, axis=1)].copy()

    # ==============================
    # 연속 투숙 계산 (기존 로직 유지)
    # ==============================
    stay_ids = []
    for crew_id, group in df.groupby('Crew ID'):
        group = group.sort_values('Arr Flt Date')
        stay_id = 0
        prev_dep = None

        for _, row in group.iterrows():
            arr = row['Arr Flt Date']
            dep = row['Dep Flt Date']

            if prev_dep is None:
                stay_id += 1
            elif (arr.date() - prev_dep.date()).days >= 1:
                stay_id += 1

            stay_ids.append(f"{crew_id}_{stay_id}")
            prev_dep = dep

    df['stay_id'] = stay_ids

    # ==============================
    # stay 집계
    # Dep_Flt: 레이오버 편수 포함된 stay면 그 편수 우선 사용
    # ==============================
    stay_records = []
    for stay_id, group in df.groupby('stay_id'):
        group = group.sort_values('Arr Flt Date')
        crew_id = group['Crew ID'].iloc[0]

        checkin  = group['Arr Flt Date'].min()
        checkout = group['Dep Flt Date'].max()
        arr_flt  = group['Arr Flt'].iloc[0]

        dep_flt_upper_grp = group['Dep Flt'].astype(str).str.strip().str.upper()
        layover_in_group  = dep_flt_upper_grp.isin(layover_flights)
        if layover_flights and layover_in_group.any():
            dep_flt = group.loc[layover_in_group, 'Dep Flt'].iloc[0]
        else:
            dep_flt = group['Dep Flt'].iloc[-1]

        stay_records.append({
            'stay_id':   stay_id,
            'Crew ID':   crew_id,
            'Crew Name': group['Crew Name'].iloc[0],
            'Rank':      group['Rank'].iloc[0],
            'Checkin':   checkin,
            'Checkout':  checkout,
            'Arr_Flt':   arr_flt,
            'Dep_Flt':   dep_flt,
        })

    stay_df = pd.DataFrame(stay_records)

    # ==============================
    # 레이오버 처리: Dep_Flt가 레이오버 편수인 stay만 Checkin+3일로 고정
    # extended_stay_ids: stay_id 단위로 추적 (같은 크루의 다른 stay는 파란색 제외)
    # ==============================
    extended_stay_ids = set()  # ✅ crew_id 대신 stay_id로 추적
    if layover_flights:
        dep_flt_upper_stay = stay_df['Dep_Flt'].astype(str).str.strip().str.upper()
        layover_stay_mask  = dep_flt_upper_stay.isin(layover_flights)
        for idx in stay_df[layover_stay_mask].index:
            checkin_date = stay_df.at[idx, 'Checkin'].date()
            stay_df.at[idx, 'Checkout'] = pd.Timestamp(checkin_date) + pd.Timedelta(days=3)
            extended_stay_ids.add(stay_df.at[idx, 'stay_id'])  # ✅ stay_id 저장

    stay_df['Nights_num'] = (stay_df['Checkout'].dt.date - stay_df['Checkin'].dt.date).apply(lambda x: x.days)
    stay_df['Nights']     = stay_df['Nights_num'].astype(str) + "박"

    stay_df = stay_df[['stay_id', 'Crew ID', 'Crew Name', 'Rank', 'Arr_Flt', 'Dep_Flt', 'Checkin', 'Checkout', 'Nights']]

    # 체크인 날짜+시간 → 도착편(같은 편수끼리 묶임) → 이름 순 정렬
    stay_df = stay_df.sort_values(['Checkin', 'Arr_Flt', 'Crew Name']).reset_index(drop=True)

    st.subheader("🗓️ Crew Hotel Stay List (필터링 완료)")

    if extended_stay_ids:
        # 레이오버 처리된 크루 이름 추출 (stay_id 기준)
        processed_names = stay_df[stay_df['stay_id'].isin(extended_stay_ids)]['Crew Name'].unique()
        st.info(f"🔄 레이오버 처리 적용됨: **{', '.join(processed_names)}** ({len(processed_names)}명) — Dep_Flt 기준 3박 고정 완료")

    # stay_id 컬럼은 내부 처리용이므로 화면 표시에서 제외
    display_df = stay_df.drop(columns=['stay_id'])
    st.dataframe(display_df, use_container_width=True)

    # ==============================
    # 엑셀 생성
    # ==============================
    output = BytesIO()
    wb_out  = Workbook()
    ws_out  = wb_out.active
    ws_out.title = "Stay List"

    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill   = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    prev_date    = None
    # 중복 Crew ID 판별: 화면에 2번 이상 나타나는 Crew ID (stay_id 기준, 다른 stay)
    duplicate_ids = stay_df['Crew ID'].astype(str)[stay_df['Crew ID'].astype(str).duplicated(keep=False)].unique()

    output_cols = ['Crew ID', 'Crew Name', 'Rank', 'Arr_Flt', 'Dep_Flt', 'Checkin', 'Checkout', 'Nights']

    for _, row in stay_df.iterrows():
        current_date = row['Checkin'].date()

        if prev_date != current_date:
            if prev_date is not None:
                ws_out.append([])

            ws_out.append([f"===== {current_date} ====="])
            cell = ws_out.cell(row=ws_out.max_row, column=1)
            cell.font      = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal="center")
            ws_out.merge_cells(
                start_row=ws_out.max_row, start_column=1,
                end_row=ws_out.max_row,   end_column=8
            )

            ws_out.append(output_cols)
            for col in range(1, 9):
                ws_out.cell(row=ws_out.max_row, column=col).fill = header_fill

        ws_out.append([
            str(row['Crew ID']), row['Crew Name'], row['Rank'],
            row['Arr_Flt'], row['Dep_Flt'],
            row['Checkin'].strftime("%Y-%m-%d %H:%M"),
            row['Checkout'].strftime("%Y-%m-%d %H:%M"),
            row['Nights']
        ])

        # ✅ 파란색: stay_id 기준으로만 레이오버 stay 행에만 적용
        # ✅ 노란색: 같은 Crew ID가 여러 stay로 등장하는 행 (레이오버 stay 우선)
        if row['stay_id'] in extended_stay_ids:
            for col in range(1, 9):
                ws_out.cell(row=ws_out.max_row, column=col).fill = blue_fill
        elif str(row['Crew ID']) in duplicate_ids:
            ws_out.cell(row=ws_out.max_row, column=1).fill = yellow_fill

        prev_date = current_date

    column_widths = {"A": 15, "B": 20, "C": 12, "D": 12, "E": 12, "F": 20, "G": 20, "H": 8}
    for col, width in column_widths.items():
        ws_out.column_dimensions[col].width = width

    wb_out.save(output)
    output.seek(0)

    st.download_button(
        label="📥 최종 엑셀 다운로드",
        data=output,
        file_name="Crew_Reservation_Final.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
